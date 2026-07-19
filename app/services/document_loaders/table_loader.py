"""CSV and XLSX loaders for exported tickets, SLOs, and dependency tables."""

from __future__ import annotations

import csv
from itertools import islice
from pathlib import Path
from typing import Any
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook

from app.services.document_loaders.base import (
    DocumentCleaningReport,
    LoadedDocument,
    base_metadata,
    filter_loaded_documents,
    normalize_text,
)

MAX_CELL_CHARS = 500
MAX_COLUMNS = 24
MAX_TABLE_ROWS = 10_000
MAX_WORKBOOK_SHEETS = 32
MAX_XLSX_UNCOMPRESSED_BYTES = 100 * 1024 * 1024
MAX_CSV_FIELD_CHARS = 10 * 1024 * 1024
MAX_TABLE_WARNINGS = 200
PRIMARY_KEY_FIELDS = ("ticket_id", "incident_id", "service_name", "error_code", "id")

csv.field_size_limit(max(csv.field_size_limit(), MAX_CSV_FIELD_CHARS))


class TableDocumentLoader:
    """Load CSV/XLSX rows as row-level auditable knowledge units."""

    loader_type = "table"
    supported_extensions = {"csv", "xlsx"}

    def load(self, path: Path) -> tuple[list[LoadedDocument], DocumentCleaningReport]:
        suffix = path.suffix.lower().removeprefix(".")
        raw_units, warnings = _load_csv(path) if suffix == "csv" else _load_xlsx(path)
        documents, report = filter_loaded_documents(
            path,
            loader_type=self.loader_type,
            raw_units=raw_units,
            min_chars=10,
        )
        report.extend_warnings(warnings)
        return documents, report


def _load_csv(path: Path) -> tuple[list[LoadedDocument], list[str]]:
    warnings: list[str] = []
    documents: list[LoadedDocument] = []
    headers, rows = _read_csv_rows(path)
    if not headers:
        _append_warning(warnings, "CSV has no header row")
        return documents, warnings
    headers = _normalize_headers(headers, warnings=warnings, sheet_name="csv")
    for index, values in enumerate(rows, 2):
        if index > MAX_TABLE_ROWS + 1:
            _append_warning(warnings, f"sheet=csv ignored rows after {MAX_TABLE_ROWS}")
            break
        documents.append(
            _row_to_document(
                path,
                row=_row_from_values(
                    headers,
                    values,
                    warnings=warnings,
                    sheet_name="csv",
                    row_number=index,
                ),
                row_number=index,
                sheet_name="csv",
                table_name=path.stem,
                warnings=warnings,
            )
        )
    if not rows:
        _append_warning(warnings, "CSV has a header row but no data rows")
    return documents, warnings


def _load_xlsx(path: Path) -> tuple[list[LoadedDocument], list[str]]:
    _validate_xlsx_archive(path)
    workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    try:
        documents: list[LoadedDocument] = []
        warnings: list[str] = []
        for sheet_index, worksheet in enumerate(workbook.worksheets, 1):
            if sheet_index > MAX_WORKBOOK_SHEETS:
                _append_warning(warnings, f"ignored sheets after {MAX_WORKBOOK_SHEETS}")
                break
            max_column = min(max(int(worksheet.max_column or 1), 1), MAX_COLUMNS + 1)
            rows = worksheet.iter_rows(
                min_row=1,
                max_row=MAX_TABLE_ROWS + 2,
                max_col=max_column,
                values_only=True,
            )
            headers = _normalize_headers(
                list(next(rows, [])),
                warnings=warnings,
                sheet_name=worksheet.title,
            )
            if not any(headers):
                _append_warning(warnings, f"sheet={worksheet.title} has no header row")
                continue
            sheet_has_rows = False
            for row_index, values in enumerate(rows, 2):
                sheet_has_rows = True
                if row_index > MAX_TABLE_ROWS + 1:
                    _append_warning(
                        warnings, f"sheet={worksheet.title} ignored rows after {MAX_TABLE_ROWS}"
                    )
                    break
                documents.append(
                    _row_to_document(
                        path,
                        row=_row_from_values(
                            headers,
                            list(values),
                            warnings=warnings,
                            sheet_name=worksheet.title,
                            row_number=row_index,
                        ),
                        row_number=row_index,
                        sheet_name=worksheet.title,
                        table_name=worksheet.title,
                        warnings=warnings,
                    )
                )
            if not sheet_has_rows:
                _append_warning(
                    warnings,
                    f"sheet={worksheet.title} has a header row but no data rows",
                )
        return documents, warnings
    finally:
        workbook.close()


def _row_to_document(
    path: Path,
    *,
    row: dict[str, Any],
    row_number: int,
    sheet_name: str,
    table_name: str,
    warnings: list[str],
) -> LoadedDocument:
    clean_items = []
    row_items = list(row.items())
    if len(row_items) > MAX_COLUMNS:
        _append_warning(
            warnings,
            f"sheet={sheet_name} row={row_number} ignored columns after {MAX_COLUMNS}",
        )
    for key, value in row_items[:MAX_COLUMNS]:
        text = normalize_text(value)
        if not text:
            continue
        if len(text) > MAX_CELL_CHARS:
            _append_warning(
                warnings,
                f"sheet={sheet_name} row={row_number} cell {key} truncated from {len(text)} chars",
            )
            text = text[:MAX_CELL_CHARS] + "...[truncated]"
        clean_items.append((str(key), text))
    if not clean_items:
        return LoadedDocument(
            content="",
            metadata=base_metadata(path, doc_type="table")
            | {
                "sheet_name": sheet_name,
                "row_number": row_number,
                "table_name": table_name,
                "primary_key": "",
            },
        )
    primary_key = _primary_key(clean_items)
    row_content = "\n".join(f"{key}: {value}" for key, value in clean_items)
    lines = [
        f"表格: {path.name}",
        f"sheet: {sheet_name}",
        f"row: {row_number}",
    ]
    if primary_key:
        lines.append(f"primary_key: {primary_key}")
    lines.append(row_content)
    return LoadedDocument(
        content="\n".join(lines),
        quality_content=row_content,
        deduplication_content=row_content,
        metadata=base_metadata(path, doc_type="table")
        | {
            "sheet_name": sheet_name,
            "row_number": row_number,
            "table_name": table_name,
            "primary_key": primary_key,
        },
    )


def _clean_header(value: Any) -> str:
    return normalize_text(value).strip().replace(" ", "_").lower()


def _normalize_headers(
    values: list[Any],
    *,
    warnings: list[str],
    sheet_name: str,
) -> list[str]:
    headers: list[str] = []
    counts: dict[str, int] = {}
    for column_index, value in enumerate(values, 1):
        base = _clean_header(value) or f"column_{column_index}"
        count = counts.get(base, 0) + 1
        counts[base] = count
        header = base if count == 1 else f"{base}_{count}"
        if not _clean_header(value):
            _append_warning(
                warnings,
                f"sheet={sheet_name} column={column_index} used generated header {header}",
            )
        elif count > 1:
            _append_warning(
                warnings,
                f"sheet={sheet_name} duplicate header {base} renamed to {header}",
            )
        headers.append(header)
    return headers


def _row_from_values(
    headers: list[str],
    values: list[Any],
    *,
    warnings: list[str],
    sheet_name: str,
    row_number: int,
) -> dict[str, Any]:
    effective_headers = list(headers)
    existing_headers = set(effective_headers)
    for column_index in range(len(effective_headers) + 1, len(values) + 1):
        header = _unique_generated_header(column_index, existing_headers)
        effective_headers.append(header)
        existing_headers.add(header)
        _append_warning(
            warnings,
            f"sheet={sheet_name} row={row_number} extra column {column_index} "
            f"used generated header {header}",
        )
    return {
        header: values[index] if index < len(values) else None
        for index, header in enumerate(effective_headers)
    }


def _unique_generated_header(column_index: int, existing_headers: set[str]) -> str:
    base = f"column_{column_index}"
    if base not in existing_headers:
        return base
    suffix = 2
    while f"{base}_{suffix}" in existing_headers:
        suffix += 1
    return f"{base}_{suffix}"


def _append_warning(warnings: list[str], warning: str) -> None:
    if len(warnings) < MAX_TABLE_WARNINGS - 1:
        warnings.append(warning)
    elif len(warnings) == MAX_TABLE_WARNINGS - 1:
        warnings.append("additional table warnings omitted")


def _primary_key(items: list[tuple[str, str]]) -> str:
    values = {key.lower(): value for key, value in items}
    for field in PRIMARY_KEY_FIELDS:
        if values.get(field):
            return f"{field}={values[field]}"
    return ""


def _read_csv_rows(path: Path) -> tuple[list[Any], list[list[Any]]]:
    """Read CSV rows with encodings common in exported Chinese ticket systems."""
    last_error: UnicodeDecodeError | None = None
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                sample = handle.read(64 * 1024)
                handle.seek(0)
                reader = csv.reader(handle, dialect=_detect_csv_dialect(sample))
                headers = list(next(reader, []))
                return headers, list(islice(reader, MAX_TABLE_ROWS + 1))
        except UnicodeDecodeError as exc:
            last_error = exc
    if last_error is not None:
        raise last_error
    return [], []


def _detect_csv_dialect(sample: str) -> type[csv.Dialect]:
    """Detect common export delimiters while retaining comma CSV as the fallback."""
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        return csv.excel


def _validate_xlsx_archive(path: Path) -> None:
    """Reject malformed or highly expanded XLSX archives before workbook parsing."""
    try:
        with ZipFile(path) as archive:
            expanded_size = sum(max(int(info.file_size), 0) for info in archive.infolist())
    except BadZipFile as exc:
        raise ValueError("XLSX 文件损坏或不是有效工作簿") from exc
    if expanded_size > MAX_XLSX_UNCOMPRESSED_BYTES:
        raise ValueError(f"XLSX 解压后内容超过限制（最大 {MAX_XLSX_UNCOMPRESSED_BYTES} 字节）")
