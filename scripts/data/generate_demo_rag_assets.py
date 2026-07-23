"""Generate realistic multi-source RAG assets for the interview golden chains."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

ROOT = Path(__file__).resolve().parents[2]
DOCS_DIR = ROOT / "docs" / "knowledge-base"
FIXTURES_DIR = ROOT / "tests" / "fixtures" / "knowledge-base"


REDIS_POSTMORTEM = {
    "title": "Redis Client Capacity Exhaustion - Sanitized Incident Review",
    "incident_id": "INC-REDIS-2026-07-A",
    "owner": "Commerce Platform SRE",
    "classification": "Internal training fixture - identifiers and traffic values sanitized",
    "reviewed_at": "2026-07-21",
    "source_reference": "INC-REDIS-2026-07-A / CR-REDIS-2026-071",
    "pages": [
        {
            "heading": "1. Executive Summary and Impact",
            "paragraphs": [
                (
                    "On 2026-07-06, order-service experienced elevated checkout failures after "
                    "a release increased retry concurrency in the promotion lookup path. Redis "
                    "client slots approached the configured maxclients ceiling, new connections "
                    "waited or failed, and retries amplified the pressure."
                ),
                (
                    "The incident affected the order submission path for 18 minutes. Peak HTTP "
                    "5xx reached 12.8%, P95 latency rose from 420 ms to 4.6 s, and approximately "
                    "7.4% of checkout attempts required a user retry. No data loss was observed."
                ),
                (
                    "This review is retrospective evidence. A current healthy Redis INFO result "
                    "only proves adapter connectivity; it does not prove that the historical "
                    "incident is still active."
                ),
            ],
            "table": [
                ["Scope", "Observed impact"],
                ["Service", "order-service / promotion lookup"],
                ["Window", "2026-07-06 10:00-10:18 UTC"],
                ["Customer effect", "Checkout timeout, 5xx, delayed order confirmation"],
                ["Peak signals", "connected_clients=9940, maxclients=10000, blocked_clients=37"],
                ["Data integrity", "No confirmed loss or duplicate order writes"],
            ],
        },
        {
            "heading": "2. Incident Timeline",
            "paragraphs": [
                (
                    "The timeline correlates deploy history, Prometheus metrics, Loki logs, and "
                    "Redis INFO captured in the same incident window. Values outside this window "
                    "were treated as supporting context rather than direct proof."
                )
            ],
            "table": [
                ["UTC", "Event and evidence"],
                ["09:10", "order-api rc1 raised promotion lookup retry count"],
                ["10:00", "HighErrorRate and RedisConnectionWait alerts began firing"],
                ["10:03", "connected_clients crossed 9,900; blocked_clients reached 37"],
                ["10:06", "Loki showed Redis timeout and connection-pool wait messages"],
                ["10:09", "Incident commander froze unrelated production changes"],
                ["10:12", "Retry reduction plan approved; canary configuration prepared"],
                ["10:15", "Canary error rate and client growth began falling"],
                ["10:18", "5xx and P95 returned below recovery thresholds"],
                ["10:28", "rc2 completed after observation and approval"],
            ],
        },
        {
            "heading": "3. Evidence Register",
            "paragraphs": [
                (
                    "The incident commander recorded each artifact with its observation window and "
                    "diagnostic role. Historical tickets were supporting context only and were not "
                    "treated as proof of the live incident."
                )
            ],
            "table": [
                ["Evidence", "Observation", "Supports", "Limit"],
                [
                    "Prometheus",
                    "Clients 9,940/10,000; blocked=37",
                    "Capacity pressure",
                    "Does not identify owner",
                ],
                [
                    "Loki",
                    "Pool acquire timeout after rc1",
                    "Application impact",
                    "Sampled error logs",
                ],
                [
                    "Redis INFO",
                    "Rejected connections increased",
                    "Server-side refusal",
                    "Point-in-time snapshot",
                ],
                [
                    "Deploy history",
                    "Retry concurrency changed at 09:10",
                    "Candidate trigger",
                    "Correlation alone",
                ],
                [
                    "Canary",
                    "Client growth and 5xx declined",
                    "Causal reversal",
                    "10% traffic scope",
                ],
            ],
        },
        {
            "heading": "4. Hypothesis Review and Root Cause",
            "paragraphs": [
                (
                    "Hypothesis A - Redis CPU saturation: rejected. Redis CPU remained below 48% "
                    "and command latency did not rise before client-slot pressure."
                ),
                (
                    "Hypothesis B - network packet loss: rejected. Node and load-balancer packet "
                    "loss stayed within baseline, and failures correlated with connection creation."
                ),
                (
                    "Hypothesis C - a single slow command: not primary. Slowlog had no command whose "
                    "timing explained the connection growth or the application pool waits."
                ),
                (
                    "Hypothesis D - client-capacity exhaustion caused by retry amplification: "
                    "confirmed. The release increased retry concurrency, idle connections were "
                    "retained longer than expected, connected_clients approached maxclients, and "
                    "the rollback of retry behavior reversed all three signals."
                ),
                (
                    "Root cause: an application retry-policy change multiplied concurrent Redis "
                    "connection demand while the pool retained idle connections. The combined load "
                    "exhausted available client capacity and produced a positive feedback loop."
                ),
                (
                    "Diagnostic conclusion: the confirmed hypothesis is supported by temporal "
                    "ordering, mechanism evidence, and causal reversal. The rejected hypotheses "
                    "remain documented to prevent future responders from repeating the same checks."
                ),
            ],
        },
        {
            "heading": "5. Response, Approval, and Follow-up",
            "paragraphs": [
                (
                    "Immediate response was evidence-first: preserve the incident window, identify "
                    "connection owners, freeze unrelated changes, and prepare a reversible canary. "
                    "The agent did not restart Redis or change maxclients automatically."
                ),
                (
                    "Approved change CR-REDIS-2026-071 authorized reducing retry count and idle-pool "
                    "retention on a 10% canary. The approver was the incident commander with Redis "
                    "service owner concurrence. Rollback criteria were any increase in checkout 5xx, "
                    "P95 above 2 s for five minutes, or replica health degradation."
                ),
                (
                    "Recovery required connected_clients below 80% of effective capacity, "
                    "blocked_clients returning to baseline, checkout 5xx below 1%, and stable "
                    "replication for 30 minutes."
                ),
                (
                    "Recovery verification was owned by Commerce Platform SRE. The incident record "
                    "was reviewed on 2026-07-21 and linked to the approval and long-term action "
                    "tickets so the resolution remains auditable."
                ),
            ],
            "table": [
                ["Record", "Change or action", "Owner / approver", "Result / due"],
                [
                    "Approval",
                    "CR-REDIS-2026-071, 10% retry-policy canary",
                    "Incident Commander + Redis Owner",
                    "Approved 10:12 UTC",
                ],
                [
                    "Action",
                    "Cap retry concurrency and add jitter",
                    "Order Platform",
                    "2026-07-20 / load test",
                ],
                [
                    "Action",
                    "Alert on effective client headroom",
                    "SRE Observability",
                    "2026-07-24 / alert drill",
                ],
                [
                    "Action",
                    "Set idle pool limits by workload",
                    "Runtime Platform",
                    "2026-07-31 / config audit",
                ],
                [
                    "Action",
                    "Quarterly capacity review",
                    "Redis Service Owner",
                    "Recurring / review record",
                ],
            ],
        },
    ],
}

MYSQL_POSTMORTEM = {
    "title": "MySQL Slow Query and Pool Waiting - Sanitized Incident Review",
    "incident_id": "INC-MYSQL-2026-07-B",
    "owner": "Payments Reliability",
    "classification": "Internal training fixture - identifiers and traffic values sanitized",
    "reviewed_at": "2026-07-21",
    "source_reference": "INC-MYSQL-2026-07-B / CR-MYSQL-2026-044",
    "pages": [
        {
            "heading": "1. Executive Summary and Impact",
            "paragraphs": [
                (
                    "On 2026-07-06, payment-service checkout latency increased after release rc3 "
                    "changed ORM eager loading for the order report query. The new query plan read "
                    "far more rows, held MySQL connections longer, and caused application "
                    "pool_waiting during peak traffic."
                ),
                (
                    "The customer-visible window lasted 19 minutes. Payment P95 rose from 610 ms to "
                    "5.2 s, timeout rate peaked at 9.6%, and 3.1% of payment attempts were retried. "
                    "Idempotency controls prevented duplicate charges."
                ),
                (
                    "This document records a historical diagnosis. Deploy correlation alone is not "
                    "root-cause proof; the release evidence was accepted only after matching the SQL "
                    "digest, connection occupancy, and incident-window latency."
                ),
            ],
            "table": [
                ["Scope", "Observed impact"],
                ["Service", "payment-service / checkout"],
                ["Window", "2026-07-06 10:05-10:24 UTC"],
                ["Peak database signals", "slow_queries=18, active_connections=188/200"],
                ["Application signal", "pool_waiting=6, checkout timeout"],
                ["Integrity", "No confirmed duplicate charge or lost payment event"],
            ],
        },
        {
            "heading": "2. Incident Timeline",
            "paragraphs": [
                (
                    "Evidence was aligned to the incident window and separated into deployment, "
                    "database, application, and customer-impact signals."
                )
            ],
            "table": [
                ["UTC", "Event and evidence"],
                ["09:42", "payment-api rc3 changed checkout query loading"],
                ["10:05", "PaymentLatencyHigh alert began firing"],
                ["10:08", "slow_queries=18 and active_connections=188/200"],
                ["10:10", "Loki recorded digest payment_report_join_v3 and pool waits"],
                ["10:13", "Read-only EXPLAIN showed high row estimate and temporary table"],
                ["10:15", "Incident commander approved disabling report feature on canary"],
                ["10:18", "Canary pool_waiting fell to zero; timeout rate declined"],
                ["10:24", "Core payment SLO recovered"],
                ["10:32", "rc4 completed with report feature disabled"],
            ],
        },
        {
            "heading": "3. Evidence Register",
            "paragraphs": [
                (
                    "Evidence was captured before any database mutation. Read-only EXPLAIN and "
                    "application metrics were used to separate a query-plan regression from host, "
                    "network, and permanent connection-leak hypotheses."
                )
            ],
            "table": [
                ["Evidence", "Observation", "Supports", "Limit"],
                [
                    "Prometheus",
                    "Connections 188/200; pool_waiting=6",
                    "Pool pressure",
                    "Does not name SQL",
                ],
                [
                    "Slow query digest",
                    "payment_report_join_v3 appeared after rc3",
                    "Query correlation",
                    "Digest omits literals",
                ],
                [
                    "Read-only EXPLAIN",
                    "Wide join and temporary table",
                    "Plan regression",
                    "Estimate, not runtime trace",
                ],
                [
                    "Deploy history",
                    "ORM eager loading changed at 09:42",
                    "Candidate trigger",
                    "Correlation alone",
                ],
                [
                    "Canary",
                    "Disabling feature cleared pool waits",
                    "Causal reversal",
                    "10% traffic scope",
                ],
            ],
        },
        {
            "heading": "4. Hypothesis Review and Root Cause",
            "paragraphs": [
                (
                    "Hypothesis A - database host CPU exhaustion: rejected. CPU increased only after "
                    "the slow digest volume rose and remained below the saturation threshold."
                ),
                (
                    "Hypothesis B - network latency between application and MySQL: rejected. TCP "
                    "connect latency and packet loss stayed near baseline."
                ),
                (
                    "Hypothesis C - connection leak: not primary. Connections returned to the pool "
                    "after requests completed; their hold time, not permanent leakage, increased."
                ),
                (
                    "Hypothesis D - rc3 query-plan regression: confirmed. The digest appeared after "
                    "deployment, EXPLAIN showed a wider join and temporary table, connection hold "
                    "time increased, and disabling the feature reversed pool_waiting."
                ),
                (
                    "Root cause: an ORM eager-loading change expanded the checkout report join. The "
                    "query held connections long enough to consume the application pool, creating "
                    "queueing and request timeouts."
                ),
                (
                    "Diagnostic conclusion: the confirmed hypothesis is supported by the new digest, "
                    "read-only plan evidence, connection hold time, and causal reversal. Rejected "
                    "hypotheses are retained as negative evidence for future incident comparison."
                ),
            ],
        },
        {
            "heading": "5. Response, Approval, and Follow-up",
            "paragraphs": [
                (
                    "The initial response captured the SQL digest, transaction state, pool metrics, "
                    "and deploy history before proposing a change. EXPLAIN was run through a "
                    "read-only path. No index, pool-size, SQL, or restart change was executed by the "
                    "agent."
                ),
                (
                    "Approved change CR-MYSQL-2026-044 disabled the report feature on a 10% canary. "
                    "The payments service owner approved the application change; the DBA approved "
                    "the later covering-index plan. Rollback criteria included increased lock wait, "
                    "replication lag above 10 s, or payment error rate above 2%."
                ),
                (
                    "Recovery required pool_waiting=0, active connections below 70% of capacity, "
                    "P95 below 1 s, no duplicate-charge signal, and 30 minutes of stable observation."
                ),
                (
                    "Recovery verification was owned by Payments Reliability and the DBA. The "
                    "incident record was reviewed on 2026-07-21 and linked to the approval, query "
                    "regression test, and covering-index action tickets."
                ),
            ],
            "table": [
                ["Record", "Change or action", "Owner / approver", "Result / due"],
                [
                    "Approval",
                    "CR-MYSQL-2026-044, disable report on 10% canary",
                    "Payments Owner + DBA",
                    "Approved 10:15 UTC",
                ],
                [
                    "Action",
                    "Add query-plan regression test",
                    "Payments",
                    "2026-07-22 / CI fixture",
                ],
                [
                    "Action",
                    "Create covering index after review",
                    "DBA",
                    "2026-07-25 / EXPLAIN + canary",
                ],
                [
                    "Action",
                    "Alert on pool wait and hold time",
                    "SRE",
                    "2026-07-24 / alert drill",
                ],
                [
                    "Action",
                    "Add release-to-digest correlation",
                    "Observability",
                    "2026-07-31 / dashboard",
                ],
            ],
        },
    ],
}

REDIS_CAPACITY_WIKI_HTML = """<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <title>Redis Capacity Runbook - Client Headroom</title>
</head>
<body>
  <h1>Redis Capacity Runbook</h1>
  <h2>Scope and ownership</h2>
  <p>This runbook applies when Redis clients are rejected, connection-pool wait rises,
  connected_clients approaches effective maxclients, or retry amplification is suspected.
  It does not cover command latency without connection pressure.</p>
  <p>Owner: Redis Service Team. Escalation: Commerce Platform SRE. Last reviewed:
  2026-07-21. Related tickets: INC-REDIS-001, INC-REDIS-009, and CR-REDIS-2026-071.</p>
  <p>This document is the Redis Capacity Wiki. It distinguishes live_info from
  incident-window evidence and records the maxclients approval boundary.</p>

  <h2>Evidence and metric queries</h2>
  <p>Capture incident-window values for connected_clients, blocked_clients,
  rejected_connections_total, application pool waiting, HTTP 5xx, and P95 latency.
  Compare connected_clients with the effective limit after operating-system file descriptor
  reservations, not only the configured maxclients value.</p>
  <pre>redis_connected_clients / redis_config_maxclients
rate(redis_rejected_connections_total[5m])
histogram_quantile(0.95, sum by (le, service) (rate(http_request_duration_seconds_bucket[5m])))</pre>

  <h2>Log patterns</h2>
  <p>Search for ERR max number of clients reached, Redis connection timeout, pool acquire
  timeout, connection refused, retry exhausted, and repeated reconnect attempts. Group by
  service, release, pod, and client library to identify the connection owner.</p>

  <h2>Decision tree</h2>
  <p>If connected_clients is below 70% and rejected connections are zero, investigate
  network latency or slow commands instead. If client headroom is below 10%, identify retry
  growth and idle-pool retention. If one release owns the growth, prepare a reversible
  application canary. If growth is broad, prepare capacity review with the Redis owner.</p>

  <h2>Safe response</h2>
  <p>First preserve Redis INFO, application pool metrics, deploy history, and logs from the
  same incident window. Then generate a change plan with expected effect, canary scope,
  approver, rollback command, and observation period. Do not restart Redis, raise maxclients,
  or resize pools automatically.</p>

  <h2>Approval and rollback conditions</h2>
  <p>Changing maxclients, operating-system file limits, Redis capacity, retry policy, or pool
  limits requires human approval and a production change window. Roll back if error rate,
  replication lag, memory pressure, or connection churn worsens during the canary.</p>

  <h2>Recovery criteria</h2>
  <p>Declare recovery only when connected_clients remains below 80% of effective capacity,
  blocked_clients and rejected connections return to baseline, customer error rate is below
  the service threshold, and the signals remain stable for at least 30 minutes.</p>

  <h2>Historical context</h2>
  <p>The sanitized golden incident recorded connected_clients=9940, maxclients=10000, and
  blocked_clients=37. Treat these values as historical evidence, not as current live_info.
  See redis_postmortem.pdf and the tickets sheet in tickets.xlsx.</p>
</body>
</html>
"""

PAYMENT_WIKI_HTML = """<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <title>Payment Runbook - MySQL Query and Pool Waiting</title>
</head>
<body>
  <h1>Payment MySQL Runbook</h1>
  <h2>Scope and ownership</h2>
  <p>Retrieval alias: Payment Runbook MySQL slow query EXPLAIN active_connections
  pool_waiting.</p>
  <p>This runbook applies when checkout latency, MySQL slow-query volume, active connection
  occupancy, or application pool_waiting rises. Owner: Payments Reliability. DBA escalation:
  Database Platform. Last reviewed: 2026-07-21. Related tickets: INC-MYSQL-014,
  INC-MYSQL-021, and CR-MYSQL-2026-044.</p>

  <h2>Evidence and metric queries</h2>
  <p>Align slow_queries, active_connections, pool_waiting, connection hold time, HTTP latency,
  timeout rate, and deploy history to the same incident window. A recent release is a
  hypothesis until a SQL digest or plan change connects it to the symptoms.</p>
  <pre>rate(mysql_global_status_slow_queries[5m])
mysql_global_status_threads_connected / mysql_global_variables_max_connections
sum by (service) (application_db_pool_waiting)
histogram_quantile(0.95, sum by (le) (rate(http_request_duration_seconds_bucket[5m])))</pre>

  <h2>Log and digest patterns</h2>
  <p>Search for checkout timeout, pool acquire timeout, lock wait timeout, deadlock,
  rows examined, temporary table, filesort, and the normalized SQL digest. Group by release,
  endpoint, tenant class, and database user.</p>

  <h2>Read-only diagnosis</h2>
  <p>Capture the normalized digest and run EXPLAIN through an approved read-only path.
  Compare estimated rows, chosen indexes, join order, temporary-table use, and lock behavior
  with the previous known-good plan. Do not run an unbounded production query to reproduce
  the issue.</p>

  <h2>Decision tree</h2>
  <p>If pool_waiting rises while connection hold time rises, prioritize slow SQL or lock
  contention. If connections rise without longer hold time, investigate a leak or pool
  sizing. If lock waits dominate, use the database lock-wait runbook. If only one release
  introduces the digest, prepare a feature rollback or canary disable plan.</p>

  <h2>Change and approval boundary</h2>
  <p>Adding an index, rewriting SQL, changing pool size, terminating sessions, restarting
  services, or disabling a production feature requires a reviewed change plan and human
  approval. The plan must include expected rows affected, lock risk, replica impact,
  rollback steps, and observation thresholds.</p>

  <h2>Rollback conditions</h2>
  <p>Roll back a canary if payment error rate exceeds 2%, lock wait rises, replica lag exceeds
  10 seconds, active connections continue growing, or the target SQL digest becomes slower.
  Prefer application feature rollback before emergency database mutation when it removes the
  triggering query safely.</p>

  <h2>Recovery criteria and history</h2>
  <p>Recovery requires pool_waiting=0, active connections below 70% of capacity, payment P95
  below 1 second, stable replication, and no duplicate-charge signal for 30 minutes. The
  sanitized incident facts were slow_queries=18, active_connections=188/200, pool_waiting=6,
  and release payment-api-2026.07.06-rc3.</p>
</body>
</html>
"""

TICKET_ROWS = [
    {
        "ticket_id": "INC-REDIS-001",
        "service_name": "order-service",
        "incident_type": "redis_maxclients",
        "impact": "Checkout 5xx peaked at 12.8% for 18 minutes",
        "root_cause": "Redis maxclients exhausted by retry storm",
        "excluded_hypotheses": "Redis CPU saturation; packet loss; single slow command",
        "decision_record": "Confirmed after temporal ordering, mechanism evidence, and canary reversal",
        "resolution": "Reduced retry amplification and raised maxclients after approval",
        "evidence": "connected_clients=9940 maxclients=10000 blocked_clients=37",
        "approval_record": "CR-REDIS-2026-071 approved by Incident Commander and Redis Owner",
        "rollback_condition": "Checkout 5xx or replication lag increases during canary",
        "owner": "Commerce Platform SRE",
        "updated_at": "2026-07-21",
    },
    {
        "ticket_id": "INC-REDIS-009",
        "service_name": "order-service",
        "incident_type": "redis_maxclients",
        "impact": "促销查询重试超时 delayed order confirmation",
        "root_cause": "促销查询重试循环 exhausted Redis client slots and maxclients headroom",
        "excluded_hypotheses": "Network loss; Redis host CPU; command latency",
        "decision_record": "Confirmed by release ownership and client-pressure reversal",
        "resolution": "Reduced retry burst and capped idle Redis pool after approval",
        "evidence": "Loki redis timeout Prometheus 5xx connected_clients near maxclients",
        "approval_record": "Order Platform and Redis Owner approved 10% canary",
        "rollback_condition": "Client churn, blocked clients, or 5xx increases",
        "owner": "Order Platform",
        "updated_at": "2026-07-21",
    },
    {
        "ticket_id": "INC-MYSQL-014",
        "service_name": "payment-service",
        "incident_type": "mysql_slow_query",
        "impact": "Payment P95 reached 5.2 seconds and timeout rate reached 9.6%",
        "root_cause": "Slow SQL held MySQL connections and caused pool waiting",
        "excluded_hypotheses": "Database host CPU; network latency; permanent connection leak",
        "decision_record": "Confirmed by digest, EXPLAIN, hold time, and canary reversal",
        "resolution": "Captured digest, added index after approval, observed P95 recovery",
        "evidence": "slow_queries=18 active_connections=188/200 pool_waiting=6",
        "approval_record": "CR-MYSQL-2026-044 approved by Payments Owner and DBA",
        "rollback_condition": "Lock wait, replica lag, or payment errors increase",
        "owner": "Payments Reliability",
        "updated_at": "2026-07-21",
    },
    {
        "ticket_id": "INC-MYSQL-021",
        "service_name": "payment-service",
        "incident_type": "mysql_pool_waiting",
        "impact": "Checkout requests queued while idempotency prevented duplicate charges",
        "root_cause": "Checkout report query caused MySQL pool_waiting after release rc3",
        "excluded_hypotheses": "Connection leak; packet loss; storage saturation",
        "decision_record": "Confirmed after feature disable cleared pool waiting",
        "resolution": "Disabled report flag, reviewed EXPLAIN, then added covering index",
        "evidence": "deploy rc3 slow query digest payment_report_join_v3 pool_waiting=6",
        "approval_record": "Feature canary approved before the later DBA index change",
        "rollback_condition": "Payment errors exceed 2% or replica lag exceeds 10 seconds",
        "owner": "Payments Reliability",
        "updated_at": "2026-07-21",
    },
    {
        "ticket_id": "INC-NET-027",
        "service_name": "gateway-service",
        "incident_type": "network_timeout",
        "impact": "Cross-zone connect latency caused intermittent upstream 504 responses",
        "root_cause": "SNAT port pressure amplified by short-lived retry connections",
        "excluded_hypotheses": "DNS resolution; TLS expiry; application first-byte latency",
        "decision_record": "Confirmed by network-stage decomposition and connection canary",
        "resolution": "Reduced retry concurrency and shifted a 10% canary after approval",
        "evidence": "connect latency retransmits conntrack and SNAT utilization aligned",
        "approval_record": "Network Platform approved route canary CR-NET-2026-027",
        "rollback_condition": "Packet loss, cross-zone cost, or downstream connections increase",
        "owner": "Network Platform",
        "updated_at": "2026-07-21",
    },
    {
        "ticket_id": "INC-TLS-011",
        "service_name": "public-api",
        "incident_type": "tls_certificate_expiry",
        "impact": "A subset of older clients failed TLS handshake at one ingress",
        "root_cause": "One ingress retained an expired certificate binding",
        "excluded_hypotheses": "Client clock skew; DNS mismatch; unsupported cipher",
        "decision_record": "Confirmed by node-specific certificate serial and binding comparison",
        "resolution": "Applied dual-certificate canary and reloaded the ingress after approval",
        "evidence": "SNI serial notAfter and ingress binding differed on one node",
        "approval_record": "Security Platform and API Owner approved CR-TLS-2026-011",
        "rollback_condition": "Handshake failures or legacy-client failures increase",
        "owner": "Security Platform",
        "updated_at": "2026-07-21",
    },
    {
        "ticket_id": "INC-DNS-019",
        "service_name": "catalog-service",
        "incident_type": "dns_resolution_failure",
        "impact": "Pods intermittently returned no such host for an internal dependency",
        "root_cause": "CoreDNS upstream timeout combined with stale negative cache",
        "excluded_hypotheses": "Authoritative record error; service endpoint failure; TCP routing",
        "decision_record": "Confirmed by resolver comparison and CoreDNS upstream timing",
        "resolution": "Canaried an upstream resolver change after preserving DNS evidence",
        "evidence": "SERVFAIL coredns latency and Pod resolver comparison",
        "approval_record": "Kubernetes Platform approved CR-DNS-2026-019",
        "rollback_condition": "NXDOMAIN, wrong-address answers, or traffic skew increases",
        "owner": "Kubernetes Platform",
        "updated_at": "2026-07-21",
    },
    {
        "ticket_id": "INC-THREAD-008",
        "service_name": "inventory-service",
        "incident_type": "thread_pool_exhaustion",
        "impact": "Request queue growth pushed inventory P99 above eight seconds",
        "root_cause": "Worker threads blocked on a slow downstream API",
        "excluded_hypotheses": "CPU saturation; database pool leak; garbage collection",
        "decision_record": "Confirmed by thread stacks and downstream span duration",
        "resolution": "Reduced downstream concurrency and enabled approved backpressure canary",
        "evidence": "thread dumps queue depth rejected tasks and downstream trace spans",
        "approval_record": "Inventory Owner approved CR-THREAD-2026-008",
        "rollback_condition": "Rejected tasks, downstream connections, or error rate increases",
        "owner": "Inventory Platform",
        "updated_at": "2026-07-21",
    },
    {
        "ticket_id": "INC-MQ-023",
        "service_name": "fulfillment-consumer",
        "incident_type": "message_queue_backlog",
        "impact": "Oldest fulfillment event age reached 26 minutes",
        "root_cause": "A poison message triggered repeated consumer retries on one partition",
        "excluded_hypotheses": "Producer surge; broker disk saturation; broad consumer shortage",
        "decision_record": "Confirmed by single-message and single-partition concentration",
        "resolution": "Quarantined the message and replayed at a capped rate after approval",
        "evidence": "single partition lag repeated message key and downstream validation error",
        "approval_record": "Messaging Platform and Fulfillment Owner approved replay",
        "rollback_condition": "Duplicate processing, rebalance, or downstream errors increase",
        "owner": "Messaging Platform",
        "updated_at": "2026-07-21",
    },
    {
        "ticket_id": "INC-K8S-031",
        "service_name": "recommendation-worker",
        "incident_type": "kubernetes_scheduling_failure",
        "impact": "New worker replicas remained Pending and reduced batch throughput",
        "root_cause": "Resource requests and zone affinity left no eligible node",
        "excluded_hypotheses": "Container image failure; application crash; PVC outage",
        "decision_record": "Confirmed by scheduler events and eligible-node calculation",
        "resolution": "Adjusted one canary workload constraint after platform approval",
        "evidence": "FailedScheduling events requests allocatable capacity and affinity",
        "approval_record": "Kubernetes Platform approved CR-K8S-2026-031",
        "rollback_condition": "Wrong-zone placement, resource contention, or instability appears",
        "owner": "Kubernetes Platform",
        "updated_at": "2026-07-21",
    },
    {
        "ticket_id": "INC-MYSQL-LOCK-017",
        "service_name": "billing-service",
        "incident_type": "mysql_lock_wait",
        "impact": "Invoice updates queued and API timeout rate reached 6.4%",
        "root_cause": "A long batch transaction blocked the online update path",
        "excluded_hypotheses": "Slow storage; connection leak; database CPU saturation",
        "decision_record": "Confirmed by lock graph, transaction age, and recovery after rollback",
        "resolution": "Paused the batch and rolled back the blocker after DBA approval",
        "evidence": "performance_schema lock graph transaction age digest and pool waiting",
        "approval_record": "DBA and Billing Owner approved CR-LOCK-2026-017",
        "rollback_condition": "Replica lag, rollback volume, or data mismatch increases",
        "owner": "Database Platform",
        "updated_at": "2026-07-21",
    },
]


def main() -> None:
    DOCS_DIR.mkdir(parents=True, exist_ok=True)
    FIXTURES_DIR.mkdir(parents=True, exist_ok=True)
    write_postmortem_pdf(DOCS_DIR / "redis_postmortem.pdf", REDIS_POSTMORTEM)
    write_postmortem_pdf(DOCS_DIR / "mysql_slow_query_postmortem.pdf", MYSQL_POSTMORTEM)
    (DOCS_DIR / "redis_capacity_wiki.html").write_text(
        REDIS_CAPACITY_WIKI_HTML, encoding="utf-8"
    )
    (DOCS_DIR / "payment_wiki.html").write_text(PAYMENT_WIKI_HTML, encoding="utf-8")
    write_tickets_csv_fixture(FIXTURES_DIR / "tickets.csv")
    write_tickets_xlsx(DOCS_DIR / "tickets.xlsx")
    print(f"Generated production RAG assets in {DOCS_DIR}")
    print(f"Generated CSV loader fixture in {FIXTURES_DIR}")


def write_tickets_csv_fixture(path: Path) -> None:
    headers = list(TICKET_ROWS[0])
    lines = [",".join(headers)]
    for row in TICKET_ROWS:
        values = [str(row[key]).replace('"', '""') for key in headers]
        lines.append(",".join(f'"{value}"' if "," in value else value for value in values))
    path.write_text("\n".join(lines) + "\n", encoding="utf-8-sig")


def write_tickets_xlsx(path: Path) -> None:
    workbook = Workbook()
    tickets = workbook.active
    tickets.title = "tickets"
    headers = list(TICKET_ROWS[0].keys())
    tickets.append(headers)
    for row in TICKET_ROWS:
        tickets.append([row[key] for key in headers])
    _format_sheet(tickets, widths=[22, 24, 26, 42, 46, 42, 46, 46, 48, 46, 44, 26, 16])

    deploys = workbook.create_sheet("deploy_history")
    deploys.append(["service_name", "version", "deployed_at", "change_summary", "risk_hint"])
    deploys.append(
        [
            "payment-service",
            "payment-api-2026.07.06-rc3",
            "2026-07-06T09:42:00Z",
            "Changed checkout order query path and ORM eager loading",
            "Correlates with slow query and pool_waiting=6",
        ]
    )
    deploys.append(
        [
            "payment-service",
            "payment-api-2026.07.06-rc4",
            "2026-07-06T10:32:00Z",
            "Disabled checkout report feature flag and prepared index change",
            "Remediation after approval; not an automatically executed Agent action",
        ]
    )
    deploys.append(
        [
            "order-service",
            "order-api-2026.07.06-rc1",
            "2026-07-06T09:10:00Z",
            "Raised Redis retry count in promotion lookup path",
            "Can amplify Redis maxclients pressure",
        ]
    )
    deploys.append(
        [
            "order-service",
            "order-api-2026.07.06-rc2",
            "2026-07-06T10:28:00Z",
            "Reduced Redis retry count and idle pool retention",
            "Remediation after approval for maxclients pressure",
        ]
    )
    _format_sheet(deploys, widths=[24, 34, 24, 54, 54])
    workbook.save(path)


def _format_sheet(sheet, *, widths: list[int]) -> None:
    """Apply a restrained audit-friendly workbook layout."""
    header_fill = PatternFill("solid", fgColor="17324D")
    header_font = Font(color="FFFFFF", bold=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 30
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.row_dimensions[row[0].row].height = 66
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def write_postmortem_pdf(path: Path, report: dict[str, object]) -> None:
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleCentered",
        parent=styles["Title"],
        alignment=TA_CENTER,
        fontName="Helvetica-Bold",
        fontSize=20,
        leading=24,
        textColor=colors.HexColor("#17324D"),
        spaceAfter=8,
    )
    subtitle_style = ParagraphStyle(
        "Subtitle",
        parent=styles["BodyText"],
        alignment=TA_CENTER,
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#52606D"),
        spaceAfter=14,
    )
    heading_style = ParagraphStyle(
        "SectionHeading",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=15,
        leading=19,
        textColor=colors.HexColor("#0B6E69"),
        spaceAfter=10,
    )
    body_style = ParagraphStyle(
        "Body",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=9.5,
        leading=14,
        textColor=colors.HexColor("#202B33"),
        spaceAfter=8,
    )

    document = SimpleDocTemplate(
        str(path),
        pagesize=A4,
        rightMargin=18 * mm,
        leftMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=17 * mm,
        title=str(report["title"]),
        author=str(report["owner"]),
        subject="Sanitized incident postmortem for RAG retrieval",
    )
    story = []
    pages = list(report["pages"])
    for page_index, page in enumerate(pages):
        if page_index == 0:
            story.append(Paragraph(str(report["title"]), title_style))
            story.append(
                Paragraph(
                    f"{report['incident_id']} | Owner: {report['owner']}<br/>"
                    f"{report['classification']}<br/>"
                    f"Reviewed: {report['reviewed_at']} | Trace: {report['source_reference']}",
                    subtitle_style,
                )
            )
        story.append(Paragraph(str(page["heading"]), heading_style))
        for paragraph in page.get("paragraphs", []):
            story.append(Paragraph(str(paragraph), body_style))
        if page.get("table"):
            raw_table = page["table"]
            column_count = len(raw_table[0])
            if column_count == 2:
                column_widths = [45 * mm, 112 * mm]
            else:
                column_widths = [34 * mm, 59 * mm, 34 * mm, 30 * mm]
            cell_style = ParagraphStyle(
                "TableCell",
                parent=body_style,
                fontSize=8.2,
                leading=10.5,
                spaceAfter=0,
            )
            header_cell_style = ParagraphStyle(
                "TableHeaderCell",
                parent=cell_style,
                fontName="Helvetica-Bold",
                textColor=colors.white,
            )
            rendered_table = [
                [
                    Paragraph(str(value), header_cell_style if row_index == 0 else cell_style)
                    for value in row
                ]
                for row_index, row in enumerate(raw_table)
            ]
            table = Table(rendered_table, repeatRows=1, colWidths=column_widths)
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#17324D")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                        ("LEADING", (0, 0), (-1, -1), 11),
                        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#A7B6C2")),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F7F8")]),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 6),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                        ("TOPPADDING", (0, 0), (-1, -1), 5),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ]
                )
            )
            story.extend([Spacer(1, 3 * mm), table])
        if page_index < len(pages) - 1:
            story.append(PageBreak())

    document.build(story, onFirstPage=_draw_page, onLaterPages=_draw_page)


def _draw_page(canvas, document) -> None:
    canvas.saveState()
    canvas.setStrokeColor(colors.HexColor("#D5DEE3"))
    canvas.line(18 * mm, 14 * mm, A4[0] - 18 * mm, 14 * mm)
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.HexColor("#657681"))
    canvas.drawString(18 * mm, 9 * mm, "AutoOnCall sanitized knowledge asset")
    canvas.drawRightString(A4[0] - 18 * mm, 9 * mm, f"Page {document.page}")
    canvas.restoreState()


if __name__ == "__main__":
    main()
