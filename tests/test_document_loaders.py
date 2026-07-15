"""Tests for multi-source RAG document loaders."""

from __future__ import annotations

from pathlib import Path
from types import SimpleNamespace
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook

from app.services.document_loaders.html_loader import HtmlDocumentLoader
from app.services.document_loaders.pdf_loader import PdfDocumentLoader
from app.services.document_loaders.plain_text import PlainTextLoader
from app.services.document_loaders.registry import document_loader_registry
from app.services.document_loaders.table_loader import TableDocumentLoader
from app.services.rag_answer_policy import ensure_citation_block
from app.services.rag_read_models import compact_retrieval_chunk
from scripts.data.generate_demo_rag_assets import main as generate_demo_rag_assets


def test_loader_registry_supports_enterprise_knowledge_formats() -> None:
    assert {
        "txt",
        "md",
        "markdown",
        "pdf",
        "html",
        "htm",
        "csv",
        "xlsx",
    }.issubset(document_loader_registry.supported_extensions)


def test_html_loader_removes_navigation_and_preserves_heading_path(tmp_path: Path) -> None:
    html = tmp_path / "wiki.html"
    html.write_text(
        """
        <html>
          <body>
            <nav>首页 导航 联系我们</nav>
            <h1>Redis 故障处理</h1>
            <h2>maxclients 耗尽</h2>
            <p>order-service 出现 Redis connection timeout，需要检查 connected_clients。</p>
            <aside>related docs noise should be removed</aside>
            <form><button>submit noisy search form</button></form>
            <footer>版权信息</footer>
            <script>alert("noise")</script>
          </body>
        </html>
        """,
        encoding="utf-8",
    )

    docs, report = HtmlDocumentLoader().load(html)

    assert report.raw_units >= 1
    assert report.indexed_units == 1
    assert docs[0].metadata["doc_type"] == "html"
    assert docs[0].metadata["heading_path"] == "Redis 故障处理 > maxclients 耗尽"
    assert "connected_clients" in docs[0].content
    assert "首页 导航" not in docs[0].content
    assert "related docs noise" not in docs[0].content
    assert "submit noisy search" not in docs[0].content
    assert "alert" not in docs[0].content


def test_html_loader_respects_declared_gb18030_encoding(tmp_path: Path) -> None:
    html = tmp_path / "wiki.html"
    content = """
        <html>
            <head><meta charset="gb18030"></head>
            <body>
              <h1>故障处理</h1>
              <p>订单服务连接超时，需要检查连接池、上游依赖状态和最近一次配置变更记录。</p>
            </body>
          </html>
    """
    html.write_bytes(content.encode("gb18030"))

    docs, report = HtmlDocumentLoader().load(html)

    assert report.indexed_units == 1
    assert docs[0].metadata["heading_path"] == "故障处理"
    assert "订单服务连接超时" in docs[0].content
    assert "连接池" in docs[0].content


def test_html_loader_limits_heading_sections(monkeypatch, tmp_path: Path) -> None:
    monkeypatch.setattr("app.services.document_loaders.html_loader.MAX_HTML_SECTIONS", 2)
    html = tmp_path / "wiki.html"
    html.write_text(
        "<h1>One</h1><p>first section with enough diagnostic context</p>"
        "<h1>Two</h1><p>second section with enough diagnostic context</p>"
        "<h1>Three</h1><p>third section with enough diagnostic context</p>",
        encoding="utf-8",
    )

    docs, report = HtmlDocumentLoader().load(html)

    assert len(docs) == 2
    assert report.raw_units == 2
    assert any("ignored sections after 2" in warning for warning in report.warnings)
    assert all("Three" not in document.content for document in docs)


def test_html_loader_does_not_duplicate_nested_semantic_blocks(tmp_path: Path) -> None:
    html = tmp_path / "wiki.html"
    html.write_text(
        "<h1>Runbook</h1>"
        "<ul><li><p>restart service after checking dependency health and error rate</p></li></ul>"
        "<pre><code>kubectl rollout restart deployment/order-service</code></pre>",
        encoding="utf-8",
    )

    docs, report = HtmlDocumentLoader().load(html)

    assert report.indexed_units == 1
    assert docs[0].content.count("restart service after checking") == 1
    assert docs[0].content.count("kubectl rollout restart") == 1


def test_plain_text_loader_preserves_markdown_code_and_list_indentation(
    tmp_path: Path,
) -> None:
    markdown = tmp_path / "runbook.md"
    markdown.write_text(
        "# Runbook\n\n"
        "```python\n"
        "if ready:\n"
        "    restart_service()\n"
        "```\n\n"
        "- parent step\n"
        "  - child step with enough diagnostic context\n",
        encoding="utf-8",
    )

    docs, report = PlainTextLoader().load(markdown)

    assert report.indexed_units == 1
    assert "    restart_service()" in docs[0].content
    assert "  - child step" in docs[0].content
    assert "```\n\n- parent step" in docs[0].content


def test_plain_text_loader_removes_utf8_bom_before_markdown_splitting(tmp_path: Path) -> None:
    markdown = tmp_path / "runbook.md"
    markdown.write_bytes(
        b"\xef\xbb\xbf# Redis\n\nRedis timeout runbook with enough diagnostic context."
    )

    docs, report = PlainTextLoader().load(markdown)

    assert report.indexed_units == 1
    assert docs[0].content.startswith("# Redis")
    assert not docs[0].content.startswith("\ufeff")


def test_plain_text_loader_drops_punctuation_only_content(tmp_path: Path) -> None:
    text_file = tmp_path / "noise.txt"
    text_file.write_text("-" * 80, encoding="utf-8")

    docs, report = PlainTextLoader().load(text_file)

    assert docs == []
    assert report.low_information_units == 1


def test_csv_loader_turns_rows_into_citable_knowledge_units(tmp_path: Path) -> None:
    csv_file = tmp_path / "tickets.csv"
    csv_file.write_text(
        "ticket_id,service_name,root_cause,resolution\n"
        "INC-REDIS-001,order-service,Redis maxclients exhausted,raise maxclients after approval\n",
        encoding="utf-8",
    )

    docs, report = TableDocumentLoader().load(csv_file)

    assert report.indexed_units == 1
    assert docs[0].metadata["doc_type"] == "table"
    assert docs[0].metadata["sheet_name"] == "csv"
    assert docs[0].metadata["row_number"] == 2
    assert docs[0].metadata["primary_key"] == "ticket_id=INC-REDIS-001"
    assert "Redis maxclients exhausted" in docs[0].content


@pytest.mark.parametrize(
    ("content", "warning"),
    [
        ("", "CSV has no header row"),
        ("ticket_id,summary\n", "CSV has a header row but no data rows"),
    ],
)
def test_csv_loader_reports_structurally_empty_files(
    tmp_path: Path,
    content: str,
    warning: str,
) -> None:
    csv_file = tmp_path / "empty.csv"
    csv_file.write_text(content, encoding="utf-8")

    docs, report = TableDocumentLoader().load(csv_file)

    assert docs == []
    assert warning in report.warnings


def test_csv_loader_drops_empty_rows_and_reports_truncation(tmp_path: Path) -> None:
    csv_file = tmp_path / "tickets.csv"
    long_resolution = "修复步骤" * 300
    csv_file.write_text(
        "ticket_id,service_name,resolution,extra1,extra2,extra3,extra4,extra5,extra6,extra7,"
        "extra8,extra9,extra10,extra11,extra12,extra13,extra14,extra15,extra16,extra17,"
        "extra18,extra19,extra20,extra21,extra22,extra23,extra24,extra25\n"
        f"INC-REDIS-001,order-service,{long_resolution},1,2,3,4,5,6,7,8,9,10,11,12,"
        "13,14,15,16,17,18,19,20,21,22,23,24,25\n"
        ",,,,,,,,,,,,,,,,,,,,,,,,,,,\n",
        encoding="utf-8",
    )

    docs, report = TableDocumentLoader().load(csv_file)

    assert len(docs) == 1
    assert report.raw_units == 2
    assert report.indexed_units == 1
    assert report.empty_units == 1
    assert any("truncated from" in warning for warning in report.warnings)
    assert any("ignored columns after" in warning for warning in report.warnings)
    assert "[truncated]" in docs[0].content


def test_xlsx_loader_preserves_sheet_and_row_locator(tmp_path: Path) -> None:
    xlsx_file = tmp_path / "service_catalog.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "services"
    sheet.append(["service_name", "owner", "dependency"])
    sheet.append(["payment-service", "payments-oncall", "mysql-payments"])
    workbook.save(xlsx_file)

    docs, report = TableDocumentLoader().load(xlsx_file)

    assert report.indexed_units == 1
    assert docs[0].metadata["sheet_name"] == "services"
    assert docs[0].metadata["row_number"] == 2
    assert docs[0].metadata["primary_key"] == "service_name=payment-service"
    assert "mysql-payments" in docs[0].content


def test_table_loader_preserves_duplicate_and_empty_headers(tmp_path: Path) -> None:
    csv_file = tmp_path / "tickets.csv"
    csv_file.write_text(
        "service_name,service_name,,root_cause\n"
        "old-service,new-service,prod,timeout with enough diagnostic context\n",
        encoding="utf-8",
    )

    docs, report = TableDocumentLoader().load(csv_file)

    assert report.indexed_units == 1
    assert "service_name: old-service" in docs[0].content
    assert "service_name_2: new-service" in docs[0].content
    assert "column_3: prod" in docs[0].content
    assert any("duplicate header" in warning for warning in report.warnings)
    assert any("generated header" in warning for warning in report.warnings)


def test_table_loader_limits_csv_rows(monkeypatch, tmp_path: Path) -> None:
    monkeypatch.setattr("app.services.document_loaders.table_loader.MAX_TABLE_ROWS", 2)
    csv_file = tmp_path / "tickets.csv"
    csv_file.write_text(
        "ticket_id,summary\n"
        "INC-1,first timeout with enough diagnostic context\n"
        "INC-2,second timeout with enough diagnostic context\n"
        "INC-3,third timeout with enough diagnostic context\n",
        encoding="utf-8",
    )

    docs, report = TableDocumentLoader().load(csv_file)

    assert len(docs) == 2
    assert any("ignored rows after 2" in warning for warning in report.warnings)


@pytest.mark.parametrize("delimiter", [";", "\t", "|"])
def test_csv_loader_detects_common_export_delimiters(
    delimiter: str,
    tmp_path: Path,
) -> None:
    csv_file = tmp_path / "tickets.csv"
    csv_file.write_text(
        delimiter.join(["ticket_id", "service_name", "summary"])
        + "\n"
        + delimiter.join(
            ["INC-1", "order-service", "timeout with enough diagnostic context"]
        )
        + "\n",
        encoding="utf-8",
    )

    docs, report = TableDocumentLoader().load(csv_file)

    assert report.indexed_units == 1
    assert docs[0].metadata["primary_key"] == "ticket_id=INC-1"
    assert "service_name: order-service" in docs[0].content


def test_table_loader_deduplicates_identical_rows_despite_row_locator(tmp_path: Path) -> None:
    csv_file = tmp_path / "tickets.csv"
    csv_file.write_text(
        "ticket_id,summary\n"
        "INC-1,timeout with enough diagnostic context\n"
        "INC-1,timeout with enough diagnostic context\n",
        encoding="utf-8",
    )

    docs, report = TableDocumentLoader().load(csv_file)

    assert len(docs) == 1
    assert docs[0].metadata["row_number"] == 2
    assert report.duplicate_units == 1


def test_table_loader_drops_rows_with_only_tiny_values(tmp_path: Path) -> None:
    csv_file = tmp_path / "tiny.csv"
    csv_file.write_text("id\nx\n", encoding="utf-8")

    docs, report = TableDocumentLoader().load(csv_file)

    assert docs == []
    assert report.low_information_units == 1


def test_csv_loader_truncates_fields_larger_than_csv_default_limit(tmp_path: Path) -> None:
    csv_file = tmp_path / "tickets.csv"
    csv_file.write_text(
        "ticket_id,summary\nINC-1," + ("x" * 140_000) + "\n",
        encoding="utf-8",
    )

    docs, report = TableDocumentLoader().load(csv_file)

    assert report.indexed_units == 1
    assert "[truncated]" in docs[0].content
    assert any("truncated from 140000 chars" in warning for warning in report.warnings)


def test_table_loader_preserves_values_beyond_declared_headers(tmp_path: Path) -> None:
    csv_file = tmp_path / "tickets.csv"
    csv_file.write_text(
        "ticket_id,summary\n"
        "INC-1,timeout with enough diagnostic context,order-service\n",
        encoding="utf-8",
    )

    docs, report = TableDocumentLoader().load(csv_file)

    assert "column_3: order-service" in docs[0].content
    assert any("extra column 3" in warning for warning in report.warnings)


def test_xlsx_loader_preserves_formula_text_without_cached_value(tmp_path: Path) -> None:
    xlsx_file = tmp_path / "thresholds.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["service_name", "threshold"])
    sheet.append(["payment-service", "=1+2"])
    workbook.save(xlsx_file)
    workbook.close()

    docs, report = TableDocumentLoader().load(xlsx_file)

    assert report.indexed_units == 1
    assert "threshold: =1+2" in docs[0].content


def test_xlsx_loader_reports_header_only_sheet(tmp_path: Path) -> None:
    xlsx_file = tmp_path / "empty.xlsx"
    workbook = Workbook()
    workbook.active.append(["service_name", "owner"])
    workbook.save(xlsx_file)
    workbook.close()

    docs, report = TableDocumentLoader().load(xlsx_file)

    assert docs == []
    assert any("has a header row but no data rows" in item for item in report.warnings)


def test_xlsx_loader_rejects_excessive_expanded_archive(monkeypatch, tmp_path: Path) -> None:
    monkeypatch.setattr(
        "app.services.document_loaders.table_loader.MAX_XLSX_UNCOMPRESSED_BYTES",
        100,
    )
    xlsx_file = tmp_path / "oversized.xlsx"
    with ZipFile(xlsx_file, "w", compression=ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", "x" * 101)

    with pytest.raises(ValueError, match="解压后内容超过限制"):
        TableDocumentLoader().load(xlsx_file)


def test_pdf_loader_keeps_page_number_and_reports_empty_pages(monkeypatch, tmp_path: Path) -> None:
    pdf_file = tmp_path / "postmortem.pdf"
    pdf_file.write_bytes(b"%PDF fake for monkeypatched reader")

    class FakePage:
        def __init__(self, text: str) -> None:
            self.text = text

        def extract_text(self) -> str:
            return self.text

    def fake_reader(_path: str) -> SimpleNamespace:
        return SimpleNamespace(
            pages=[
                FakePage("Redis maxclients 故障复盘，connected_clients 达到 9940。"),
                FakePage(""),
            ]
        )

    monkeypatch.setattr("app.services.document_loaders.pdf_loader.PdfReader", fake_reader)

    docs, report = PdfDocumentLoader().load(pdf_file)

    assert report.raw_units == 2
    assert report.indexed_units == 1
    assert report.empty_units == 1
    assert docs[0].metadata["doc_type"] == "pdf"
    assert docs[0].metadata["page_number"] == 1
    assert "connected_clients" in docs[0].content


def test_pdf_loader_limits_page_count(monkeypatch, tmp_path: Path) -> None:
    pdf_file = tmp_path / "postmortem.pdf"
    pdf_file.write_bytes(b"%PDF fake for monkeypatched reader")
    monkeypatch.setattr("app.services.document_loaders.pdf_loader.MAX_PDF_PAGES", 2)

    class FakePage:
        def extract_text(self) -> str:
            return "Redis timeout diagnostic context with enough information."

    def fake_reader(_path: str) -> SimpleNamespace:
        return SimpleNamespace(pages=[FakePage(), FakePage(), FakePage()])

    monkeypatch.setattr("app.services.document_loaders.pdf_loader.PdfReader", fake_reader)

    docs, report = PdfDocumentLoader().load(pdf_file)

    assert len(docs) == 1
    assert report.raw_units == 2
    assert report.duplicate_units == 1
    assert any("ignored pages after 2" in warning for warning in report.warnings)


def test_pdf_loader_keeps_good_pages_when_one_page_extraction_fails(
    monkeypatch,
    tmp_path: Path,
) -> None:
    pdf_file = tmp_path / "partial.pdf"
    pdf_file.write_bytes(b"%PDF fake for monkeypatched reader")

    class GoodPage:
        def extract_text(self) -> str:
            return "Redis timeout diagnostic content with enough information."

    class BrokenPage:
        def extract_text(self) -> str:
            raise ValueError("bad stream at C:/private/partial.pdf")

    def fake_reader(_path: str) -> SimpleNamespace:
        return SimpleNamespace(pages=[GoodPage(), BrokenPage()])

    monkeypatch.setattr("app.services.document_loaders.pdf_loader.PdfReader", fake_reader)

    docs, report = PdfDocumentLoader().load(pdf_file)

    assert len(docs) == 1
    assert docs[0].metadata["page_number"] == 1
    assert report.raw_units == 2
    assert report.empty_units == 1
    assert any("page 2 text extraction failed" in warning for warning in report.warnings)
    assert "private" not in " ".join(report.warnings)


def test_citation_block_includes_pdf_and_table_locators() -> None:
    answer = "基于知识库，Redis 连接数接近上限。"
    rendered = ensure_citation_block(
        answer,
        [
            {
                "source_file": "redis_postmortem.pdf",
                "chunk_id": "redis_postmortem.pdf#0001",
                "page_number": 3,
                "score": 0.12,
            },
            {
                "source_file": "tickets.xlsx",
                "chunk_id": "tickets.xlsx#0002",
                "sheet_name": "incidents",
                "row_number": 12,
                "primary_key": "ticket_id=INC-REDIS-001",
                "score": 0.2,
            },
        ],
    )

    assert "page_number: 3" in rendered
    assert "sheet_name: incidents" in rendered
    assert "row_number: 12" in rendered
    assert "primary_key: ticket_id=INC-REDIS-001" in rendered


def test_compact_retrieval_chunk_exposes_source_specific_locators() -> None:
    compact = compact_retrieval_chunk(
        {
            "source_file": "tickets.xlsx",
            "chunk_id": "tickets.xlsx#0002",
            "metadata": {
                "doc_type": "table",
                "sheet_name": "incidents",
                "row_number": 12,
                "primary_key": "ticket_id=INC-REDIS-001",
            },
        }
    )

    assert compact["doc_type"] == "table"
    assert compact["sheet_name"] == "incidents"
    assert compact["row_number"] == 12
    assert compact["primary_key"] == "ticket_id=INC-REDIS-001"


def test_generated_demo_rag_assets_are_loader_readable(
    tmp_path: Path,
    monkeypatch,
) -> None:
    monkeypatch.setattr(
        "scripts.data.generate_demo_rag_assets.DOCS_DIR",
        tmp_path,
    )
    generate_demo_rag_assets()

    expected = {
        "redis_postmortem.pdf": "pdf",
        "mysql_slow_query_postmortem.pdf": "pdf",
        "redis_capacity_wiki.html": "html",
        "payment_wiki.html": "html",
        "tickets.csv": "table",
        "tickets.xlsx": "table",
    }
    for file_name, loader_type in expected.items():
        path = tmp_path / file_name
        loader = document_loader_registry.get_loader(path)
        docs, report = loader.load(path)

        assert loader.loader_type == loader_type
        assert docs
        assert report.indexed_units >= 1
